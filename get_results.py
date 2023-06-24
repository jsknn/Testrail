from testrail import *
import config
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
run_id = input("Enter the test run id:")
client = APIClient(config.url)
client.user = config.username
client.password = config.password
run = client.send_get('get_run/'+run_id)
defect = client.send_get('get_results_for_run/'+run_id+'&status_id=7,6,5,4,2')
Tests =  client.send_get('get_tests/'+run_id+'&status_id=7,6,5,4,2')
Test_id =  client.send_get('/get_test/9993188')
Section = client.send_get('/get_cases/'+config.project)

Title=[]
for i in Tests:
    Titles = i.get('title')
    Title.append(Titles)
    ##print (Titles) 

section=[]
for t in Tests :
    for s in Section:
        x = t.get('title')==s.get('title')
        if x is True:
           b = s.get('section_id')
           if b==981:
                b= "General"
                section.append(b)
                ##print(b)
           elif b==982:
                b= "General"
                section.append(b)
                ##print(b)
           elif b==982:
                b= "Load Url"
                section.append(b)
                ##print(b)
           elif b==1322:
                b= "Av Sync"
                section.append(b)
                ##print(b)
           elif b==984:
                b= "Set Js Value"
                section.append(b)
                ##print(b)
           elif b==985:
                b= "Widevine/DRM"
                section.append(b)
                ##print(b)
           elif b==986:
                b= "Evaluate JS"
                section.append(b)
                ##print(b)
           elif b==987:
                b= "Resize"
                section.append(b)
                ##print(b)
           elif b==989:
                b= "Set Webview Client"
                section.append(b)
                ##print(b)
           elif b==990:
                b= "Keyboard"
                section.append(b)
                ##print(b)
           elif b==991:
                b= "JS Bridge"
                section.append(b)
                ##print(b)
           elif b==992:
                b= "Audio"
                section.append(b)
                ##print(b)
           elif b==995:
                b= "Stop Loading"
                section.append(b)
                ##print(b)
           elif b==996:
                b= "Clear History"
                section.append(b)
                ##print(b)
           elif b==997:
                b= "Clear Cache"
                section.append(b)
                ##print(b)
           elif b==998:
                b= "Onload Started"
                section.append(b)
                ##print(b)
           elif b==999:
                b= "Onload Finished"
                section.append(b)
                ##print(b)
           elif b==1000:
                b= "Set UserAgent String"
                section.append(b)
                ##print(b)
           elif b==1001:
                b= "Set MixedContentMode"
                section.append(b)
                ##print(b)
           elif b==1002:
                b= "Set AllowFileAccess"
                section.append(b)
                ##print(b)
           elif b==1003:
                b= "Negative Scenario"
                section.append(b)
                ##print(b)
           elif b==1004:
                b= "Cookies Validation"
                section.append(b)
                ##print(b)
           elif b==1005:
                b= "Media Auto Gesture"
                section.append(b)
                ##print(b)
           elif b==1007:
                b= "Set devTools"
                section.append(b)
                ##print(b)
           elif b==1009:
                b= "Multiple Webview Instances"
                section.append(b)
                ##print(b)
           elif b==1126:
                b= "Security"
                section.append(b)
                ##print(b)
           elif b==7959:
                b= "Multiprocess and Sandbox"
                section.append(b)
                ##print(b)
           elif b==1913:
                b= "Video/Media Types"
                section.append(b)
                ##print(b)
           elif b==993:
                b= "App Lanuch Performance Test"
                section.append(b)
                ##print(b)
           elif b==3390:
                b= "Cool App Launch/Static Content"
                section.append(b)
                ##print(b)
           elif b==3391:
                b= "Cool App Launch/top Websites"
                section.append(b)
                ##print(b)
           elif b==12946:
                b= "Cold App Lunch"
                section.append(b)
                ##print(b)
           elif b==3392:
                b= "Memory Static Content"
                section.append(b)
                ##print(b)
           elif b==3393:
                b= "Memory Video/Scrolling Popular website"
                section.append(b)
                ##print(b)
           elif b==3394:
                b= "Memory Popular Website"
                section.append(b)
                ##print(b)
           elif b==1310:
                b= "Fluidity Fps"
                section.append(b)
                ##print(b)
           elif b==11169:
                b= "Thermal Tests"
                section.append(b)
                ##print(b)
           elif b==2381:
                b= "OnTrim memory Testing"
                section.append(b)
                ##print(b)
           elif b==9617:
                b= "Stress Testing"
                section.append(b)
                ##print(b)
           elif b==9842:
                b= "VS(VegaScript) Webview Testapp"
                section.append(b)
                ##print(b)
           elif b==15808:
                b= "VS memory Test"
                section.append(b)
                ##print(b)
           elif b==15809:
                b= "Vs App Launch Tests"
                section.append(b)
                ##print(b)
           elif b==15810:
                b= "VS Fluidity Tests"
                section.append(b)
                ##print(b)
           elif b==10651:
                b= "Error Code"
                section.append(b)
                ##print(b)
           elif b==11361:
                b= "Metrics"
                section.append(b)
                ##print(b)
           elif b==13164:
                b= "TV"
                section.append(b)
                ##print(b)
           elif b==8750:
                b= "Remote Navigation Testing"
                section.append(b)
                ##print(b)
           elif b==8860:
                b= "Functional"
                section.append(b)
                ##print(b)
           elif b==9357:
                b= "Dependency"
                section.append(b)
                ##print(b)
           elif b==3126:
                b= "I18N"
                section.append(b)
                ##print(b)
           elif b==1315:
                b= "L10N"
                section.append(b)
                ##print(b)
           elif b==1239:
                b= "Redbull"
                section.append(b)
                ##print(b)
           elif b==6327:
                b= "Tubi"
                section.append(b)
                ##print(b)
           elif b==2445:
                b= "VoiceView"
                section.append(b)
                ##print(b)
           elif b==2448:
                b= "Screen Magnifier"
                section.append(b)
                ##print(b)
           elif b==2450:
                b= "Basic Gestures"
                section.append(b)
                ##print(b)
           elif b==2468:
                b= "Color Conversion"
                section.append(b)
                ##print(b)
           elif b==2469:
                b= "Closed Captions"
                section.append(b)
                ##print(b)
           elif b==3102:
                b= "Sanity"
                section.append(b)
                ##print(b)
           elif b==983:
                b= "OOBE Test"
                section.append(b)
                ##print(b)
           elif b==3359:
                b= "Captive Portal"
                section.append(b)
                ##print(b)
           elif b==3360:
                b= "AmazonKids FreeTime"
                section.append(b)
                ##print(b)
           elif b==9522:
                b= "OZ"
                section.append(b)
                ##print(b)
           elif b==9593:
                b= "MapWebviewClient"
                section.append(b)
                ##print(b)
           elif b==9618:
                b= "APL"
                section.append(b)
                ##print(b)
           elif b==9638:
                b= "CS(Customer Service) App"
                section.append(b)
                ##print(b)
           elif b==9639:
                b= "Amazon Photos"
                section.append(b)
                ##print(b)
           elif b==9847:
                b= "System Team"
                section.append(b)
                ##print(b)
           elif b==10546:
                b= "Settings"
                section.append(b)
                ##print(b)
           else:
                section.append(b)
                ##print(b)
                   
##print ('**********Priority**********')   
priority=[]
for i in Tests:
    Priority =i.get('priority_id')
    if Priority==5:
        Priority = "Ship Stopper"
        priority.append(Priority)
    elif Priority==3:
        Priority = "High"
        priority.append(Priority)
    elif Priority==4:
        Priority = "Critical"
        priority.append(Priority)
    elif Priority==2:
        Priority = "Medium"
        priority.append(Priority)
    else:
         priority.append(Priority)
    ##print(Priority)
##print ('**********Defects**********')
Defects=[]
for i in Tests:
    for d in defect:
        x = i.get('id')==d.get('test_id')
        if x is True:
           c=d.get('defects')
           Defects.append(c)
           ##print(c)  

##print('**********Status**********')
status=[]
for i in Tests:
    Status = i.get('status_id')
    if Status ==5:
        Status = "FAILED"
        status.append(Status)
    elif Status==2:
        Status = "BLOCKED"
        status.append(Status)
    elif Status==6:
        Status = "PASSED CAUTION" 
        status.append(Status)
    elif Status==7:
        Status = "SKIPPED"  
        status.append(Status)
    elif Status==3:
        Status = "UNTESTED"
        status.append(Status)
    elif Status==4:
         Status= "RETEST"
         status.append(Status)
    else:
         status.append(Status)
    


#print(len(section))
#print(len(priority))
#print(len(Defects))
#print (len(Title))
#print(len(status))

#Create workbook object
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title='Results'

#Generate data

#Add titles in the first row of each column
sheet.cell(row=1, column=1).value='Title'
sheet.cell(row=1, column=1).fill = PatternFill(start_color="00008B", end_color="00008B",fill_type = "solid")
sheet.cell(row=1, column=1).font = Font(name="Calibri", size=12,bold=True,color="FFFFFF")
sheet.cell(row=1, column=2).value='Section'
sheet.cell(row=1, column=2).font = Font(name="Calibri", size=12,bold=True,color="FFFFFF")
sheet.cell(row=1, column=2).fill = PatternFill(start_color="00008B", end_color="00008B",fill_type = "solid")
sheet.cell(row=1, column=3).value='Priority'
sheet.cell(row=1, column=3).font = Font(name="Calibri", size=12,bold=True,color="FFFFFF")
sheet.cell(row=1, column=3).fill = PatternFill(start_color="00008B", end_color="00008B",fill_type = "solid")
sheet.cell(row=1, column=4).value='Defects'
sheet.cell(row=1, column=4).font = Font(name="Calibri", size=12,bold=True,color="FFFFFF")
sheet.cell(row=1, column=4).fill = PatternFill(start_color="00008B", end_color="00008B",fill_type = "solid")
sheet.cell(row=1, column=5).value='Status'
sheet.cell(row=1, column=5).font = Font(name="Calibri", size=12,bold=True,color="FFFFFF")
sheet.cell(row=1, column=5).fill = PatternFill(start_color="00008B", end_color="00008B",fill_type = "solid")


#Loop to set the value of each cell
for i in range(0, len(Title)):
    sheet.cell(row=i+2, column=1).value=Title[i]
    sheet.cell(row=i+2, column=2).value=section[i]
    sheet.cell(row=i+2, column=3).value=priority[i]
    sheet.cell(row=i+2, column=4).hyperlink='https://issues.labcollab.net/browse/'+str(Defects[i]!=None)
    sheet.cell(row=i+2, column=4).value=Defects[i]
    sheet.cell(row=i+2, column=4).style = "Hyperlink"
    sheet.cell(row=i+2, column=5).value=status[i]

#Finally, save the file and give it a name
wb.save(run_id+'.xlsx')

print('The results are exported to excel sheet name: '+run_id+'.xlsx')


